"""Additional tests for office document extraction to improve coverage."""

from pathlib import Path

import pytest
from docx import Document as DocxDocument
from odf import text as odf_text
from odf.opendocument import OpenDocumentText, OpenDocumentSpreadsheet, OpenDocumentPresentation
from odf.table import Table, TableRow, TableCell
from odf.draw import Page as DrawPage
from openpyxl import Workbook

from mcp_agent_rag.rag.extractor import DocumentExtractor


@pytest.fixture
def sample_xlsx_file(temp_dir: Path) -> Path:
    """Create a sample XLSX file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Add some data
    ws['A1'] = "Header1"
    ws['B1'] = "Header2"
    ws['A2'] = "Data1"
    ws['B2'] = "Data2"
    ws['A3'] = 123
    ws['B3'] = 456
    
    # Add another sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2['A1'] = "More data"
    
    xlsx_path = temp_dir / "sample.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


@pytest.fixture
def sample_odt_file(temp_dir: Path) -> Path:
    """Create a sample ODT file."""
    doc = OpenDocumentText()
    
    # Add paragraphs
    p1 = odf_text.P(text="First paragraph in ODT")
    doc.text.addElement(p1)
    
    p2 = odf_text.P(text="Second paragraph in ODT")
    doc.text.addElement(p2)
    
    odt_path = temp_dir / "sample.odt"
    doc.save(str(odt_path))
    return odt_path


@pytest.fixture
def sample_ods_file(temp_dir: Path) -> Path:
    """Create a sample ODS file."""
    doc = OpenDocumentSpreadsheet()
    
    # Create a table
    table = Table(name="TestTable")
    
    # Add rows
    for i in range(3):
        tr = TableRow()
        for j in range(2):
            tc = TableCell()
            p = odf_text.P(text=f"Cell{i}{j}")
            tc.addElement(p)
            tr.addElement(tc)
        table.addElement(tr)
    
    doc.spreadsheet.addElement(table)
    
    ods_path = temp_dir / "sample.ods"
    doc.save(str(ods_path))
    return ods_path


@pytest.fixture
def sample_odp_file(temp_dir: Path) -> Path:
    """Create a sample ODP file."""
    doc = OpenDocumentPresentation()
    
    # Create a page (slide) with required attributes
    page = DrawPage(name="page1", stylename="dp1")
    
    # Add text to the page
    p = odf_text.P(text="Slide content")
    page.addElement(p)
    
    doc.presentation.addElement(page)
    
    odp_path = temp_dir / "sample.odp"
    doc.save(str(odp_path))
    return odp_path


def test_extract_xlsx(sample_xlsx_file):
    """Test extracting text from XLSX file."""
    text = DocumentExtractor.extract_text(sample_xlsx_file)
    
    assert text is not None
    assert "TestSheet" in text
    assert "Header1" in text
    assert "Header2" in text
    assert "Data1" in text
    assert "Data2" in text
    assert "123" in text
    assert "456" in text
    assert "Sheet2" in text
    assert "More data" in text


def test_extract_odt(sample_odt_file):
    """Test extracting text from ODT file."""
    text = DocumentExtractor.extract_text(sample_odt_file)
    
    assert text is not None
    assert "First paragraph in ODT" in text
    assert "Second paragraph in ODT" in text


def test_extract_ods(sample_ods_file):
    """Test extracting text from ODS file."""
    text = DocumentExtractor.extract_text(sample_ods_file)
    
    assert text is not None
    assert "TestTable" in text
    assert "Cell00" in text
    assert "Cell11" in text
    assert "Cell22" in text


def test_extract_odp(sample_odp_file):
    """Test extracting text from ODP file."""
    text = DocumentExtractor.extract_text(sample_odp_file)
    
    assert text is not None
    assert "Slide 1" in text
    assert "Slide content" in text


def test_extract_xlsx_with_empty_cells(temp_dir: Path):
    """Test extracting XLSX with empty cells."""
    wb = Workbook()
    ws = wb.active
    
    # Add data with empty cells
    ws['A1'] = "First"
    ws['B1'] = None  # Empty cell
    ws['C1'] = "Third"
    ws['A2'] = None
    ws['B2'] = "Middle"
    ws['C2'] = None
    
    xlsx_path = temp_dir / "empty_cells.xlsx"
    wb.save(xlsx_path)
    
    text = DocumentExtractor.extract_text(xlsx_path)
    
    assert text is not None
    assert "First" in text
    assert "Third" in text
    assert "Middle" in text


def test_extract_ods_with_unnamed_table(temp_dir: Path):
    """Test extracting ODS with unnamed table."""
    doc = OpenDocumentSpreadsheet()
    
    # Create a table without name attribute
    table = Table()
    
    # Add a row
    tr = TableRow()
    tc = TableCell()
    p = odf_text.P(text="Test")
    tc.addElement(p)
    tr.addElement(tc)
    table.addElement(tr)
    
    doc.spreadsheet.addElement(table)
    
    ods_path = temp_dir / "unnamed_table.ods"
    doc.save(str(ods_path))
    
    text = DocumentExtractor.extract_text(ods_path)
    
    assert text is not None
    assert "Test" in text


def test_extract_odp_with_empty_paragraphs(temp_dir: Path):
    """Test extracting ODP with empty paragraphs."""
    doc = OpenDocumentPresentation()
    
    page = DrawPage(name="page1", stylename="dp1")
    
    # Add empty paragraph
    p1 = odf_text.P(text="")
    page.addElement(p1)
    
    # Add paragraph with content
    p2 = odf_text.P(text="Non-empty")
    page.addElement(p2)
    
    # Add another empty paragraph
    p3 = odf_text.P(text="   ")
    page.addElement(p3)
    
    doc.presentation.addElement(page)
    
    odp_path = temp_dir / "empty_paras.odp"
    doc.save(str(odp_path))
    
    text = DocumentExtractor.extract_text(odp_path)
    
    assert text is not None
    assert "Non-empty" in text


def test_extract_ods_with_empty_rows(temp_dir: Path):
    """Test extracting ODS with empty rows."""
    doc = OpenDocumentSpreadsheet()
    
    table = Table(name="EmptyRowsTable")
    
    # Add a row with content
    tr1 = TableRow()
    tc1 = TableCell()
    p1 = odf_text.P(text="Content")
    tc1.addElement(p1)
    tr1.addElement(tc1)
    table.addElement(tr1)
    
    # Add an empty row
    tr2 = TableRow()
    tc2 = TableCell()
    p2 = odf_text.P(text="")
    tc2.addElement(p2)
    tr2.addElement(tc2)
    table.addElement(tr2)
    
    doc.spreadsheet.addElement(table)
    
    ods_path = temp_dir / "empty_rows.ods"
    doc.save(str(ods_path))
    
    text = DocumentExtractor.extract_text(ods_path)
    
    assert text is not None
    assert "Content" in text
    assert "EmptyRowsTable" in text


def test_extract_xlsx_multiple_sheets(temp_dir: Path):
    """Test extracting XLSX with multiple sheets."""
    wb = Workbook()
    
    # First sheet
    ws1 = wb.active
    ws1.title = "First"
    ws1['A1'] = "First sheet data"
    
    # Second sheet
    ws2 = wb.create_sheet("Second")
    ws2['A1'] = "Second sheet data"
    
    # Third sheet
    ws3 = wb.create_sheet("Third")
    ws3['A1'] = "Third sheet data"
    
    xlsx_path = temp_dir / "multi_sheet.xlsx"
    wb.save(xlsx_path)
    
    text = DocumentExtractor.extract_text(xlsx_path)
    
    assert text is not None
    assert "First" in text
    assert "Second" in text
    assert "Third" in text
    assert "First sheet data" in text
    assert "Second sheet data" in text
    assert "Third sheet data" in text


def test_extract_odt_with_multiple_paragraphs(temp_dir: Path):
    """Test extracting ODT with multiple paragraphs."""
    doc = OpenDocumentText()
    
    # Add multiple paragraphs
    for i in range(5):
        p = odf_text.P(text=f"Paragraph {i+1}")
        doc.text.addElement(p)
    
    odt_path = temp_dir / "multi_para.odt"
    doc.save(str(odt_path))
    
    text = DocumentExtractor.extract_text(odt_path)
    
    assert text is not None
    for i in range(5):
        assert f"Paragraph {i+1}" in text


def test_extract_odp_multiple_slides(temp_dir: Path):
    """Test extracting ODP with multiple slides."""
    doc = OpenDocumentPresentation()
    
    # Add multiple pages
    for i in range(3):
        page = DrawPage(name=f"page{i+1}", stylename="dp1")
        p = odf_text.P(text=f"Slide {i+1} content")
        page.addElement(p)
        doc.presentation.addElement(page)
    
    odp_path = temp_dir / "multi_slide.odp"
    doc.save(str(odp_path))
    
    text = DocumentExtractor.extract_text(odp_path)
    
    assert text is not None
    for i in range(3):
        assert f"Slide {i+1}" in text
        assert f"Slide {i+1} content" in text
